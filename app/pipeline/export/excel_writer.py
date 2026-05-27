"""
Author: sizhchan
Org: dgaudit
Version: v0.1
Date: 2026-05-27
"""

"""
Excel (.xlsx) export module.

Writes all extracted tables into a single .xlsx workbook,
with each table in its own worksheet. Sheet names are derived
from table titles (truncated to 31 chars, OpenPyXL limit).
"""

import logging
import os

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter

from app.pipeline.ir.types import TableData, TableCell

logger = logging.getLogger(__name__)

# Excel limits
MAX_SHEET_NAME_LEN = 31
# Characters not allowed in sheet names
SHEET_NAME_FORBIDDEN = set("[]:*?/\\")

# Styling
HEADER_FONT = Font(name="SimHei", bold=True, size=11)
CELL_FONT = Font(name="SimSun", size=10)
HEADER_FILL = PatternFill(start_color="D9E2F3", end_color="D9E2F3", fill_type="solid")
THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)
CENTER_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT_ALIGN = Alignment(horizontal="left", vertical="center", wrap_text=True)


def write_excel(tables: list[TableData], output_path: str) -> str:
    """
    Write a list of extracted tables into a single .xlsx workbook.

    Args:
        tables: List of TableData objects (cross-page merged, if applicable).
        output_path: Path for the output .xlsx file.

    Returns:
        The output path (confirmed written).
    """
    wb = Workbook()
    # Remove the default empty sheet
    wb.remove(wb.active)

    if not tables:
        # Create an empty sheet so the file isn't corrupted
        ws = wb.create_sheet(title="无表格")
        ws["A1"] = "未检测到表格"
        wb.save(output_path)
        logger.info("Excel written with no tables: %s", output_path)
        return output_path

    for idx, table in enumerate(tables):
        sheet_name = _make_sheet_name(table, idx)
        ws = wb.create_sheet(title=sheet_name)
        _write_table_to_sheet(ws, table)

    wb.save(output_path)
    logger.info("Excel written: %s (%d sheets)", output_path, len(tables))
    return output_path


def _make_sheet_name(table: TableData, fallback_index: int) -> str:
    """
    Generate a valid Excel sheet name from table title.

    Rules:
    - Use table.title if available and valid.
    - Fall back to "Sheet_N" if no title.
    - Truncate to 31 characters.
    - Remove forbidden characters.
    """
    name = table.sheet_name or table.title or f"Sheet_{fallback_index + 1}"

    # Remove forbidden characters
    name = "".join(c for c in name if c not in SHEET_NAME_FORBIDDEN)

    # Truncate
    if len(name) > MAX_SHEET_NAME_LEN:
        # Try to cut at a natural boundary
        name = name[:MAX_SHEET_NAME_LEN - 3] + "..."

    # Ensure non-empty
    if not name.strip():
        name = f"Table_{fallback_index + 1}"

    return name


def _write_table_to_sheet(ws, table: TableData) -> None:
    """
    Write a single TableData into a worksheet with formatting.

    The layout:
      Row 1: Optional title (merged across all columns)
      Row 2: Column headers
      Row 3+: Data rows
    """
    if not table.rows:
        ws["A1"] = "(空表格)"
        return

    # Determine max columns
    max_cols = max(len(row) for row in table.rows)
    if table.headers:
        max_cols = max(max_cols, len(table.headers))

    start_row = 1

    # Write title row (if present and not already in headers)
    if table.title and table.title not in table.headers:
        ws.merge_cells(
            start_row=start_row,
            start_column=1,
            end_row=start_row,
            end_column=max_cols,
        )
        cell = ws.cell(row=start_row, column=1, value=table.title)
        cell.font = Font(name="SimHei", bold=True, size=12)
        cell.alignment = CENTER_ALIGN
        start_row += 1

    # Write headers
    if table.headers:
        for col_idx, header in enumerate(table.headers):
            if col_idx >= max_cols:
                break
            cell = ws.cell(row=start_row, column=col_idx + 1, value=header)
            cell.font = HEADER_FONT
            cell.fill = HEADER_FILL
            cell.alignment = CENTER_ALIGN
            cell.border = THIN_BORDER
        start_row += 1

    # Write data rows
    for r_idx, row_cells in enumerate(table.rows):
        row_num = start_row + r_idx
        for c_idx, cell_data in enumerate(row_cells):
            if c_idx >= max_cols:
                break
            cell = ws.cell(row=row_num, column=c_idx + 1, value=cell_data.text)
            cell.font = CELL_FONT
            cell.alignment = LEFT_ALIGN
            cell.border = THIN_BORDER

    # Adjust column widths (approximate)
    for col_idx in range(max_cols):
        col_letter = get_column_letter(col_idx + 1)
        # Set a sensible default width
        max_width = 12
        for row in ws.iter_rows(
            min_col=col_idx + 1,
            max_col=col_idx + 1,
            values_only=True,
        ):
            for cell_val in row:
                if cell_val:
                    # Rough CJK width estimate: 2 chars per "unit"
                    text_len = len(str(cell_val))
                    est_width = min(text_len * 1.2 + 2, 60)
                    max_width = max(max_width, est_width)
        ws.column_dimensions[col_letter].width = max_width
